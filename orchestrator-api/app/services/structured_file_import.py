from __future__ import annotations

import csv
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
import math
import re
from pathlib import Path
from typing import Any

from app.schemas.dataset_rebuild import StructuredImportProfile, SUPPORTED_STRUCTURED_EXTENSIONS


MAX_MATERIALIZED_ROWS = 10_000
MAX_XLSX_SHEETS = 50
SAFE_IDENTIFIER_RE = re.compile(r"[^a-z0-9_]+")
INTEGER_RE = re.compile(r"[+-]?(?:0|[1-9][0-9]*)\Z")
DECIMAL_RE = re.compile(r"[+-]?(?:0|[1-9][0-9]*)\.[0-9]+\Z")


class StructuredImportError(ValueError):
    def __init__(self, code: str) -> None:
        super().__init__(code)
        self.code = code


def import_structured_file(path: Path, *, relative_path: str, permission_scope_key: str) -> list[StructuredImportProfile]:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_STRUCTURED_EXTENSIONS:
        raise StructuredImportError("unsupported_structured_type")
    if suffix == ".csv":
        rows = _read_csv(path)
        return [_profile(relative_path, None, 0, rows, permission_scope_key)]
    return _read_xlsx(path, relative_path, permission_scope_key)


def safe_column_name(header: str) -> str:
    cleaned = SAFE_IDENTIFIER_RE.sub("_", header.strip().lower()).strip("_")
    return cleaned or "column"


def _read_csv(path: Path) -> list[dict[str, Any]]:
    try:
        with path.open(newline="", encoding="utf-8-sig") as handle:
            return list(csv.DictReader(handle))[:MAX_MATERIALIZED_ROWS]
    except Exception as exc:
        raise StructuredImportError("csv_import_failed") from exc


def _read_xlsx(path: Path, relative_path: str, permission_scope_key: str) -> list[StructuredImportProfile]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        profiles = []
        for sheet_index, sheet in enumerate(workbook.worksheets[:MAX_XLSX_SHEETS]):
            rows_iter = sheet.iter_rows(values_only=True)
            headers = next(rows_iter, None)
            if headers is None:
                continue
            keys = [str(value or "").strip() for value in headers]
            rows = [dict(zip(keys, values)) for values in rows_iter][:MAX_MATERIALIZED_ROWS]
            profiles.append(_profile(relative_path, sheet.title, sheet_index, rows, permission_scope_key, headers=keys))
        return profiles
    except Exception as exc:
        raise StructuredImportError("xlsx_import_failed") from exc


def _profile(
    relative_path: str,
    sheet_name: str | None,
    sheet_index: int,
    rows: list[dict[str, Any]],
    permission_scope_key: str,
    *,
    headers: list[str] | None = None,
) -> StructuredImportProfile:
    headers = headers or list(rows[0].keys() if rows else [])
    if not headers or any(not header for header in headers):
        raise StructuredImportError("invalid_structured_headers")
    safe_names = [safe_column_name(header) for header in headers]
    if len(set(headers)) != len(headers) or len(set(safe_names)) != len(safe_names):
        raise StructuredImportError("duplicate_structured_headers")
    stem = safe_column_name(Path(relative_path).stem)
    sheet_part = f"_{safe_column_name(sheet_name)}" if sheet_name else ""
    resource_key = f"structured:{permission_scope_key}:{stem}{sheet_part}"
    data_types = [_infer_column_type([row.get(header) for row in rows]) for header in headers]
    columns = [
        {"column_name": safe_name, "data_type": data_type, "safe_description": header, "ordinal_position": index + 1}
        for index, (header, safe_name, data_type) in enumerate(zip(headers, safe_names, data_types))
    ]
    normalized_rows = [
        {
            safe_name: _convert_value(row.get(header), data_type)
            for header, safe_name, data_type in zip(headers, safe_names, data_types)
        }
        for row in rows
    ]
    return StructuredImportProfile(
        resource_key=resource_key,
        runtime_relation_name=safe_column_name(resource_key.replace(":", "_")),
        display_name=Path(relative_path).stem if sheet_name is None else f"{Path(relative_path).stem} - {sheet_name}",
        columns=columns,
        rows=normalized_rows,
        metadata={
            "safe_location_path": relative_path,
            "sheet_name": sheet_name,
            "sheet_index": sheet_index,
            "row_count": len(normalized_rows),
            "column_count": len(columns),
        },
    )


def _infer_column_type(values: list[Any]) -> str:
    kinds = {_value_kind(value) for value in values}
    kinds.discard("null")
    if not kinds:
        return "text"
    if kinds == {"boolean"}:
        return "boolean"
    if kinds == {"integer"}:
        return "integer"
    if kinds <= {"integer", "decimal"}:
        return "decimal"
    return "text"


def _value_kind(value: Any) -> str:
    if value is None or isinstance(value, str) and not value.strip():
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, Decimal):
        return "decimal" if value.is_finite() else "text"
    if isinstance(value, float):
        return "decimal" if math.isfinite(value) else "text"
    if isinstance(value, (datetime, date, time)):
        return "text"
    if not isinstance(value, str) or value != value.strip():
        return "text"
    if value.casefold() in {"true", "false"}:
        return "boolean"
    if INTEGER_RE.fullmatch(value):
        return "integer"
    if DECIMAL_RE.fullmatch(value):
        try:
            return "decimal" if Decimal(value).is_finite() else "text"
        except InvalidOperation:
            return "text"
    return "text"


def _convert_value(value: Any, data_type: str) -> Any:
    if _value_kind(value) == "null":
        return None
    if data_type == "boolean":
        return value if isinstance(value, bool) else value.casefold() == "true"
    if data_type == "integer":
        return int(value)
    if data_type == "decimal":
        return value if isinstance(value, Decimal) else Decimal(str(value))
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)
